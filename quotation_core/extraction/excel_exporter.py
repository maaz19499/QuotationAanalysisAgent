import io
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def clean_description(desc: str) -> str:
    """Implement rule 6: Strip leading 'TDL - ' and trailing period."""
    if not desc:
        return ""
    
    cleaned = desc
    if cleaned.startswith("TDL - "):
        cleaned = cleaned[len("TDL - "):]
    
    if cleaned.endswith("."):
        cleaned = cleaned[:-1]
        
    return cleaned


def get_package(line_number: int) -> str:
    """Implement rule 4: Package grouping derived from line_number."""
    if 1 <= line_number <= 5:
        return "GIRLS BUILDING PART 1"
    elif 6 <= line_number <= 19:
        return "GIRLS BUILDING PART 2"
    elif 20 <= line_number <= 34:
        return "BOYS BUILDING"
    return "UNKNOWN PACKAGE"


def get_type_code(line_number: int) -> str:
    """Implement rule 5: TYPE codes derived from line_number."""
    type_codes = {
        1: "D", 2: "D1", 3: "A", 4: "F", 5: "J",
        6: "L", 7: "L1", 8: "A1", 9: "K", 10: "G",
        11: "G1", 12: "B", 13: "K1", 14: "E", 15: "E1",
        16: "C", 17: "I", 18: "F1", 19: "J1", 20: "L2",
        21: "L3", 22: "A2", 23: "K2", 24: "G2", 25: "G3",
        26: "B1", 27: "K3", 28: "E2", 29: "E3", 30: "C1",
        31: "I1", 32: "M", 33: "H", 34: "F2"
    }
    return type_codes.get(line_number, "")


def generate_crm_pre_qt_excel(json_data: dict[str, Any]) -> bytes:
    """
    Populate the CRM-PRE-QT Excel template from a structured JSON quotation object.
    Implementation based on exact rules provided in the guide.
    """
    # Load template workbook
    # We resolve the template relative to the project root
    project_root = Path(__file__).resolve().parent.parent.parent
    template_path = project_root / "output_template" / "Quotation Template.xlsx"
    
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found at {template_path}")

    wb = load_workbook(template_path)
    
    if "PRE-QT" not in wb.sheetnames:
        raise ValueError("Sheet 'PRE-QT' not found in template")
        
    sheet = wb["PRE-QT"]

    currency = json_data.get("currency", "")
    line_items = json_data.get("line_items", [])

    excel_row = 2

    # Provide mapping from 1-based numerical index to column letter for clarity, or just use 1-based numbers
    # For openpyxl, cell(row=x, column=y) is 1-based for both.
    
    # Pre-calculate column indices for 1-based indexing in openpyxl
    col = {
        "A": 1, "B": 2, "C": 3, "D": 4, "E": 5, "F": 6, "G": 7, "H": 8, "I": 9, "J": 10,
        "K": 11, "L": 12, "M": 13, "N": 14, "O": 15, "P": 16, "Q": 17, "R": 18, "S": 19, "T": 20,
        "U": 21, "V": 22, "W": 23, "X": 24, "Y": 25, "Z": 26, "AA": 27, "AB": 28, "AC": 29, "AD": 30,
        "AE": 31, "AF": 32, "AG": 33, "AH": 34, "AI": 35, "AJ": 36, "AK": 37, "AL": 38, "AM": 39, "AN": 40,
        "AO": 41, "AP": 42, "AQ": 43, "AR": 44, "AS": 45, "AT": 46
    }
    
    for item in line_items:
        r = excel_row
        ln = item.get("line_number", excel_row - 1)
        
        # --- STATIC values ---
        sheet.cell(row=r, column=col["B"], value="M")
        sheet.cell(row=r, column=col["D"], value="Miscellaneous")
        sheet.cell(row=r, column=col["E"], value="Miscellaneous")
        sheet.cell(row=r, column=col["I"], value=1)
        sheet.cell(row=r, column=col["J"], value=0)
        sheet.cell(row=r, column=col["Q"], value="LOCAL")
        sheet.cell(row=r, column=col["R"], value="OTHERS")
        sheet.cell(row=r, column=col["V"], value="PC")
        sheet.cell(row=r, column=col["AF"], value=0)
        sheet.cell(row=r, column=col["AG"], value=0)
        sheet.cell(row=r, column=col["AH"], value=0)
        sheet.cell(row=r, column=col["AI"], value=0)
        sheet.cell(row=r, column=col["AP"], value=0)

        # --- JSON values ---
        sheet.cell(row=r, column=col["C"], value=get_package(ln))
        sheet.cell(row=r, column=col["F"], value=get_type_code(ln))
        sheet.cell(row=r, column=col["G"], value=currency)
        sheet.cell(row=r, column=col["O"], value=item.get("product_code", ""))
        sheet.cell(row=r, column=col["P"], value=clean_description(item.get("description", "")))
        sheet.cell(row=r, column=col["W"], value=currency)
        sheet.cell(row=r, column=col["X"], value=item.get("quantity", 0))
        sheet.cell(row=r, column=col["Y"], value=item.get("unit_price", 0.0))
        sheet.cell(row=r, column=col["AA"], value=currency)
        sheet.cell(row=r, column=col["AT"], value=currency)

        # --- Formulas ---
        sheet.cell(row=r, column=col["H"], value=f"=V{r}")
        sheet.cell(row=r, column=col["Z"], value=f"=Y{r}*X{r}")
        sheet.cell(row=r, column=col["AB"], value=f"=Y{r}*AJ{r}")
        sheet.cell(row=r, column=col["AC"], value=f"=AB{r}*X{r}")
        sheet.cell(row=r, column=col["AJ"], value=f'=IF(W{r}="USD",3.77,IF(W{r}="EUR",4.5,IF(W{r}="GBP",5.15,IF(W{r}="AED",1.03,IF(W{r}="AUD",3,IF(W{r}="SAR",1,0))))))')
        sheet.cell(row=r, column=col["AL"], value=f"=Y{r}*(100-AE{r})/100*(1+AF{r}/100)*(1+AG{r}/100)*(1+AH{r}/100)*(1+AI{r}/100)")
        sheet.cell(row=r, column=col["AM"], value=f"=AL{r}*X{r}")
        sheet.cell(row=r, column=col["AN"], value=f"=AL{r}*AJ{r}")
        sheet.cell(row=r, column=col["AO"], value=f"=AN{r}*X{r}*I{r}")
        sheet.cell(row=r, column=col["AS"], value=f"=I{r}*AR{r}*X{r}")

        # EMPTY columns (not written to):
        # A, K, L, M, N, S, T, U, AD, AE, AK, AQ, AR, AU-AZ
        
        excel_row += 1

    # Save to bytes buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.read()
