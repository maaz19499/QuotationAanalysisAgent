"""Excel exporter — populates the CRM PRE-QT template from extracted data."""
import io
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from quotation_extraction.core.logging_config import get_logger

logger = get_logger(__name__)


def clean_description(desc: str) -> str:
    """Strip common supplier prefixes and trailing punctuation."""
    if not desc:
        return ""
    cleaned = desc.strip()
    cleaned = re.sub(r"^[A-Z]{2,5}\s*-\s*", "", cleaned)
    if cleaned.endswith("."):
        cleaned = cleaned[:-1]
    return cleaned.strip()


def _resolve_package(item: dict[str, Any]) -> str:
    if item.get("_project_name"):
        return str(item["_project_name"])
    item_type = item.get("item_type", "")
    if item_type == "alternative":
        return "ALTERNATIVES"
    elif item_type == "accessory":
        return "ACCESSORIES"
    return ""


def _resolve_type_code(item: dict[str, Any]) -> str:
    if item.get("item_number"):
        return str(item["item_number"])
    if item.get("model_number"):
        return str(item["model_number"])
    if item.get("item_code"):
        return str(item["item_code"])
    return ""


def generate_crm_pre_qt_excel(json_data: dict[str, Any]) -> bytes:
    """Populate the CRM-PRE-QT Excel template from a structured JSON quotation."""
    project_root = Path(__file__).resolve().parent.parent.parent
    template_path = project_root / "output_template" / "Quotation Template.xlsx"

    if not template_path.exists():
        raise FileNotFoundError(f"Template not found at {template_path}")

    wb = load_workbook(template_path)
    if "PRE-QT" not in wb.sheetnames:
        raise ValueError("Sheet 'PRE-QT' not found in template")

    sheet = wb["PRE-QT"]

    currency = json_data.get("currency", "")
    line_items = json_data.get("line_items", [])
    supplier_name = json_data.get("supplier_name", "")

    project = json_data.get("project")
    project_name = project.get("object_name", "") if isinstance(project, dict) else ""

    col = {
        "A": 1, "B": 2, "C": 3, "D": 4, "E": 5, "F": 6, "G": 7, "H": 8,
        "I": 9, "J": 10, "K": 11, "L": 12, "M": 13, "N": 14, "O": 15,
        "P": 16, "Q": 17, "R": 18, "S": 19, "T": 20, "U": 21, "V": 22,
        "W": 23, "X": 24, "Y": 25, "Z": 26, "AA": 27, "AB": 28, "AC": 29,
        "AD": 30, "AE": 31, "AF": 32, "AG": 33, "AH": 34, "AI": 35,
        "AJ": 36, "AK": 37, "AL": 38, "AM": 39, "AN": 40, "AO": 41,
        "AP": 42, "AQ": 43, "AR": 44, "AS": 45, "AT": 46,
    }

    excel_row = 2

    for item in line_items:
        r = excel_row
        item["_project_name"] = project_name

        product_code = item.get("item_code") or item.get("product_code", "")
        description = clean_description(item.get("description", ""))
        quantity = item.get("quantity") or 0
        unit_price = item.get("unit_price") or 0.0
        uom = item.get("unit_of_measure") or "PC"

        # Static defaults
        sheet.cell(row=r, column=col["B"], value="M")
        sheet.cell(row=r, column=col["D"], value="Miscellaneous")
        sheet.cell(row=r, column=col["E"], value="Miscellaneous")
        sheet.cell(row=r, column=col["I"], value=1)
        sheet.cell(row=r, column=col["J"], value=0)
        sheet.cell(row=r, column=col["Q"], value="LOCAL")
        sheet.cell(row=r, column=col["R"], value="OTHERS")
        sheet.cell(row=r, column=col["AF"], value=0)
        sheet.cell(row=r, column=col["AG"], value=0)
        sheet.cell(row=r, column=col["AH"], value=0)
        sheet.cell(row=r, column=col["AI"], value=0)
        sheet.cell(row=r, column=col["AP"], value=0)

        # Values
        sheet.cell(row=r, column=col["A"], value=item.get("line_number", excel_row - 1))
        sheet.cell(row=r, column=col["C"], value=_resolve_package(item))
        sheet.cell(row=r, column=col["F"], value=_resolve_type_code(item))
        sheet.cell(row=r, column=col["G"], value=currency)
        sheet.cell(row=r, column=col["O"], value=product_code)
        sheet.cell(row=r, column=col["P"], value=description)
        sheet.cell(row=r, column=col["V"], value=uom)
        sheet.cell(row=r, column=col["W"], value=currency)
        sheet.cell(row=r, column=col["X"], value=quantity)
        sheet.cell(row=r, column=col["Y"], value=unit_price)
        sheet.cell(row=r, column=col["AA"], value=currency)
        sheet.cell(row=r, column=col["AT"], value=currency)

        # Formulas
        sheet.cell(row=r, column=col["H"], value=f"=V{r}")
        sheet.cell(row=r, column=col["Z"], value=f"=Y{r}*X{r}")
        sheet.cell(row=r, column=col["AB"], value=f"=Y{r}*AJ{r}")
        sheet.cell(row=r, column=col["AC"], value=f"=AB{r}*X{r}")
        sheet.cell(
            row=r, column=col["AJ"],
            value=(
                f'=IF(W{r}="USD",3.77,'
                f'IF(W{r}="EUR",4.5,'
                f'IF(W{r}="GBP",5.15,'
                f'IF(W{r}="AED",1.03,'
                f'IF(W{r}="AUD",3,'
                f'IF(W{r}="SAR",1,0))))))'
            ),
        )
        sheet.cell(
            row=r, column=col["AL"],
            value=f"=Y{r}*(100-AE{r})/100*(1+AF{r}/100)*(1+AG{r}/100)*(1+AH{r}/100)*(1+AI{r}/100)",
        )
        sheet.cell(row=r, column=col["AM"], value=f"=AL{r}*X{r}")
        sheet.cell(row=r, column=col["AN"], value=f"=AL{r}*AJ{r}")
        sheet.cell(row=r, column=col["AO"], value=f"=AN{r}*X{r}*I{r}")
        sheet.cell(row=r, column=col["AS"], value=f"=I{r}*AR{r}*X{r}")

        del item["_project_name"]
        excel_row += 1

    logger.info("excel_export_completed", rows_written=excel_row - 2, supplier=supplier_name)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
